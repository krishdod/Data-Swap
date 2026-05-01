from __future__ import annotations

import csv
from pathlib import Path
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


def _cell_to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _safe_sheet_name(wb: Workbook, preferred: str | None = None) -> str:
    if preferred and preferred in wb.sheetnames:
        return preferred
    return wb.sheetnames[0]


def _is_csv(path: str) -> bool:
    return Path(path).suffix.lower() == ".csv"


def _csv_rows(path: str) -> Any:
    # utf-8-sig handles files saved by Excel with BOM.
    with open(path, "r", encoding="utf-8-sig", newline="", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            yield [_cell_to_str(v) for v in row]


def _trim_source_headers(raw_headers: list[str]) -> list[str]:
    # Keep columns up to last non-empty header (common spreadsheet pattern).
    last_idx = 0
    for i, h in enumerate(raw_headers, start=1):
        if h != "":
            last_idx = i
    headers = raw_headers[:last_idx] if last_idx else []
    headers = [h for h in headers if h != ""]
    return headers


def read_headers_xlsx(path: str, sheet_name: str | None = None) -> list[str]:
    if _is_csv(path):
        first = next(_csv_rows(path), [])
        return [h for h in first if h != ""]

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[_safe_sheet_name(wb, sheet_name)]
    headers: list[str] = []
    for cell in ws[1]:
        headers.append(_cell_to_str(cell.value))
    wb.close()
    headers = [h for h in headers if h != ""]
    return headers


@dataclass(frozen=True)
class ColumnProfile:
    header: str
    samples: list[str]


@dataclass(frozen=True)
class SourcePreview:
    headers: list[str]
    rows: list[list[str]]
    row_count: int
    profiles: list[ColumnProfile]


def read_source_preview_xlsx(
    path: str,
    sheet_name: str | None = None,
    preview_rows: int = 20,
    sample_limit_per_col: int = 50,
) -> SourcePreview:
    if _is_csv(path):
        it = _csv_rows(path)
        raw_headers = next(it, [])
        headers = _trim_source_headers(raw_headers)
        preview: list[list[str]] = []
        samples: list[list[str]] = [[] for _ in range(len(headers))]
        row_count = 0
        for row in it:
            row_count += 1
            row_str = row[: len(headers)]
            if len(row_str) < len(headers):
                row_str += [""] * (len(headers) - len(row_str))
            if len(preview) < preview_rows:
                preview.append(row_str)
            for ci, val in enumerate(row_str):
                if val != "" and len(samples[ci]) < sample_limit_per_col:
                    samples[ci].append(val)
        profiles = [ColumnProfile(header=headers[i], samples=samples[i]) for i in range(len(headers))]
        return SourcePreview(headers=headers, rows=preview, row_count=row_count, profiles=profiles)

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[_safe_sheet_name(wb, sheet_name)]

    raw_headers = [_cell_to_str(c.value) for c in ws[1]]
    headers = _trim_source_headers(raw_headers)

    preview: list[list[str]] = []
    samples: list[list[str]] = [[] for _ in range(len(headers))]

    row_count = 0
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        row_count += 1
        row_str = [_cell_to_str(v) for v in row]
        if len(preview) < preview_rows:
            preview.append(row_str)
        for ci, val in enumerate(row_str):
            if val != "" and len(samples[ci]) < sample_limit_per_col:
                samples[ci].append(val)

    wb.close()

    profiles = [ColumnProfile(header=headers[i], samples=samples[i]) for i in range(len(headers))]
    return SourcePreview(headers=headers, rows=preview, row_count=row_count, profiles=profiles)


def write_filled_template(
    template_path: str,
    source_path: str,
    mapping: dict[str, dict[str, str]],
    *,
    template_sheet_name: str | None = None,
    source_sheet_name: str | None = None,
    include_audit_sheets: bool = False,
    source_row_count_hint: int = 0,
) -> bytes:
    """
    mapping format per target header:
      { "type": "source"|"constant"|"blank", "value": <source header or constant> }
    """
    # CSV inputs/outputs and very large row counts use streaming mode for scale.
    # This keeps output deterministic while scaling to big files.
    if _is_csv(template_path) or _is_csv(source_path) or source_row_count_hint >= 50_000:
        return _write_filled_template_streaming(
            template_path=template_path,
            source_path=source_path,
            mapping=mapping,
            template_sheet_name=template_sheet_name,
            source_sheet_name=source_sheet_name,
            include_audit_sheets=include_audit_sheets,
        )

    twb = load_workbook(filename=template_path, read_only=False, data_only=False)
    tws = twb[_safe_sheet_name(twb, template_sheet_name)]

    # Read template headers from row 1 (keep empty cells as empty; we only map non-empty headers).
    template_headers_full = [_cell_to_str(c.value) for c in tws[1]]
    template_headers = [h for h in template_headers_full if h != ""]

    swb = load_workbook(filename=source_path, read_only=True, data_only=True)
    sws = swb[_safe_sheet_name(swb, source_sheet_name)]

    source_headers_raw = [_cell_to_str(c.value) for c in sws[1]]
    last_idx = 0
    for i, h in enumerate(source_headers_raw, start=1):
        if h != "":
            last_idx = i
    source_headers_raw = source_headers_raw[:last_idx] if last_idx else []
    source_headers = [h for h in source_headers_raw if h != ""]

    source_index = {h: i for i, h in enumerate(source_headers)}  # 0-based within trimmed cols

    # Validate mapping covers all template headers.
    missing = [h for h in template_headers if h not in mapping]
    if missing:
        raise ValueError(f"Missing mappings for template headers: {missing}")

    # Build column order in template based on non-empty headers positions
    template_col_positions: list[tuple[int, str]] = []
    for col_idx_1based, cell in enumerate(tws[1], start=1):
        h = _cell_to_str(cell.value)
        if h != "":
            template_col_positions.append((col_idx_1based, h))

    out_row = 2
    for src_row in sws.iter_rows(min_row=2, max_col=len(source_headers), values_only=True):
        # For each template column, write deterministically.
        for col_idx_1based, th in template_col_positions:
            spec = mapping[th]
            mtype = spec.get("type", "")
            val: Any = None
            if mtype == "source":
                src_h = spec.get("value", "")
                if src_h not in source_index:
                    raise ValueError(f"Mapped source header not found: {src_h}")
                v = src_row[source_index[src_h]] if source_index[src_h] < len(src_row) else None
                # Preserve empties as true blanks in Excel.
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    val = None
                else:
                    val = v
            elif mtype == "constant":
                c = spec.get("value", "")
                val = None if (c is None or str(c) == "") else c
            elif mtype == "blank":
                val = None
            else:
                raise ValueError(f"Invalid mapping type for '{th}': {mtype}")

            tws.cell(row=out_row, column=col_idx_1based).value = val

        out_row += 1

    # Always remove these from the output unless explicitly requested.
    if "Mapping" in twb.sheetnames:
        del twb["Mapping"]
    if "Errors" in twb.sheetnames:
        del twb["Errors"]

    if include_audit_sheets:
        # Add audit sheets back (explicitly)

        ws_map = twb.create_sheet("Mapping")
        ws_map.append(["Template Header", "Mapping Type", "Source/Constant Value"])
        for th in template_headers:
            spec = mapping[th]
            ws_map.append([th, spec.get("type", ""), spec.get("value", "")])

        # We don't compute row-level errors yet (strict mapping gate prevents most issues),
        # but keep the sheet for future validations (types, required columns, etc.).
        ws_err = twb.create_sheet("Errors")
        ws_err.append(["Row", "Template Header", "Issue", "Value"])

    from io import BytesIO

    bio = BytesIO()
    twb.save(bio)
    bio.seek(0)
    swb.close()
    twb.close()
    return bio.read()


def _write_filled_template_streaming(
    template_path: str,
    source_path: str,
    mapping: dict[str, dict[str, str]],
    *,
    template_sheet_name: str | None = None,
    source_sheet_name: str | None = None,
    include_audit_sheets: bool = False,
) -> bytes:
    from io import BytesIO
    from openpyxl import Workbook

    twb = None
    swb = None

    if _is_csv(template_path):
        template_title = "Sheet1"
        template_headers_full = next(_csv_rows(template_path), [])
    else:
        twb = load_workbook(filename=template_path, read_only=True, data_only=True)
        tws = twb[_safe_sheet_name(twb, template_sheet_name)]
        template_title = tws.title
        template_headers_full = [_cell_to_str(c.value) for c in tws[1]]
    template_headers = [h for h in template_headers_full if h != ""]

    source_rows: Any
    if _is_csv(source_path):
        src_iter = _csv_rows(source_path)
        source_headers_raw = next(src_iter, [])
        source_rows = src_iter
    else:
        swb = load_workbook(filename=source_path, read_only=True, data_only=True)
        sws = swb[_safe_sheet_name(swb, source_sheet_name)]
        source_headers_raw = [_cell_to_str(c.value) for c in sws[1]]
        source_rows = sws.iter_rows(min_row=2, max_col=len(source_headers_raw), values_only=True)

    source_headers = _trim_source_headers(source_headers_raw)
    source_index = {h: i for i, h in enumerate(source_headers)}

    missing = [h for h in template_headers if h not in mapping]
    if missing:
        raise ValueError(f"Missing mappings for template headers: {missing}")

    template_col_positions: list[tuple[int, str]] = []
    for col_idx_1based, h in enumerate(template_headers_full, start=1):
        if h != "":
            template_col_positions.append((col_idx_1based, h))

    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet(title=template_title)

    # Keep template row-1 structure exactly (including blank header slots).
    out_ws.append(template_headers_full)

    target_width = len(template_headers_full)
    for src_row_raw in source_rows:
        src_row = [_cell_to_str(v) for v in src_row_raw[: len(source_headers)]]
        if len(src_row) < len(source_headers):
            src_row += [""] * (len(source_headers) - len(src_row))
        out_row: list[Any] = [None] * target_width
        for col_idx_1based, th in template_col_positions:
            spec = mapping[th]
            mtype = spec.get("type", "")
            val: Any = None
            if mtype == "source":
                src_h = spec.get("value", "")
                if src_h not in source_index:
                    raise ValueError(f"Mapped source header not found: {src_h}")
                v = src_row[source_index[src_h]] if source_index[src_h] < len(src_row) else ""
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    val = None
                else:
                    val = v
            elif mtype == "constant":
                c = spec.get("value", "")
                val = None if (c is None or str(c) == "") else c
            elif mtype == "blank":
                val = None
            else:
                raise ValueError(f"Invalid mapping type for '{th}': {mtype}")

            out_row[col_idx_1based - 1] = val
        out_ws.append(out_row)

    if include_audit_sheets:
        ws_map = out_wb.create_sheet("Mapping")
        ws_map.append(["Template Header", "Mapping Type", "Source/Constant Value"])
        for th in template_headers:
            spec = mapping[th]
            ws_map.append([th, spec.get("type", ""), spec.get("value", "")])

        ws_err = out_wb.create_sheet("Errors")
        ws_err.append(["Row", "Template Header", "Issue", "Value"])

    bio = BytesIO()
    out_wb.save(bio)
    bio.seek(0)
    if swb is not None:
        swb.close()
    if twb is not None:
        twb.close()
    return bio.read()

